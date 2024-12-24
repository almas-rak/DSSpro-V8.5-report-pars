import pandas as pd
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from datetime import datetime

lic_date = datetime(year=2024, month=12, day=1)
current_date = datetime.now().date()


# Функция для проверки формата отметок
def check_format(input_string):
    pattern1 = r'^\d{2}:\d{2}-Н$'
    pattern2 = r'^Н-\d{2}:\d{2}$'
    return bool(re.match(pattern1, input_string) or re.match(pattern2, input_string))


# Запуск графического интерфейса для выбора файлов
Tk().withdraw()
file_path = askopenfilename(title="Выберите файл Excel для обработки", filetypes=[("Excel files", "*.xlsx")])
if not file_path:
    print("Выбор файла отменён.")
    exit()

# Чтение данных из выбранного файла
data = pd.read_excel(file_path)

# Задаём список ожидаемых столбцов
expected_columns = ['Date', 'Name', 'Person Group', 'Check-in Time',
                    'Check-out Time']

# Проверяем, что все столбцы из списка присутствуют в data
if set(expected_columns).issubset(data.columns):
    print("Все ожидаемые столбцы присутствуют.")
else:
    # Определяем отсутствующие столбцы
    missing_columns = set(expected_columns) - set(data.columns)
    # Отображаем сообщение об ошибке
    messagebox.showerror("Ошибка",
                         f"Файл не соответствует требованиям. Отсутствуют следующие столбцы: {', '.join(missing_columns)}")
    exit()

# Удаление лишних столбцов
data = data.drop(columns=['ID', 'Week', 'Away Period', 'Work Hours', 'Overtime Duration'], errors='ignore')
data['Date'] = pd.to_datetime(data['Date'], errors='coerce')

# Выборка по группам и сортировка
grouped_data = data.groupby('Person Group').apply(lambda x: x.sort_values(by=['Date', 'Name'])).reset_index(drop=True)

# Создание пустого DataFrame для результата
result = pd.DataFrame()
unique_names = data['Name'].unique()
unique_dates = sorted(data['Date'].dropna().unique())
group_names = []

first_group = True

# Проходим по каждой группе в сгруппированных данных
for group_name, group in grouped_data.groupby('Person Group'):
    group_names.append(group_name)
    # Если это не первая группа, добавляем пустую строку
    if not first_group:
        # Создаём DataFrame с одной пустой строкой и добавляем его в результат
        result = pd.concat([result, pd.DataFrame([[""]], columns=["Имя"])], ignore_index=True)

    # Добавляем название группы в колонку "Имя"
    result = pd.concat([result, pd.DataFrame([[group_name]], columns=["Имя"])], ignore_index=True)

    # Получаем уникальные имена из текущей группы и добавляем их в результат
    names_df = pd.DataFrame(group['Name'].unique(), columns=["Имя"])
    result = pd.concat([result, names_df], ignore_index=True)

    # Устанавливаем флаг, чтобы после первой группы добавлялась пустая строка
    first_group = False

# Создаём столбцы с объектами Timestamp
for date in unique_dates:
    result[date] = ""

for _, row in data.iterrows():
    name = row['Name']
    date = row['Date']
    check_in = row['Check-in Time'] if pd.notna(row['Check-in Time']) else "Н"
    check_out = row['Check-out Time'] if pd.notna(row['Check-out Time']) else "Н"
    result.loc[result['Имя'] == name, date] = f"{check_in}-{check_out}"

# Создаем Excel-файл
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "События за месяц"

# Заполнение заголовков и проверка выходных
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
for col_num, header in enumerate(result.columns, 1):
    cell = ws.cell(row=1, column=col_num)
    if isinstance(header, pd.Timestamp):
        cell.value = header.strftime('%d.%m')
        if header.weekday() in [5, 6]:
            cell.fill = PatternFill(start_color="819830", end_color="819830", fill_type="solid")
        else:
            cell.fill = header_fill
    else:
        cell.value = header
        cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

# Шаблоны цветов
fill_gray = PatternFill(start_color="93C6C6", end_color="93C6C6", fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
fill_yellow = PatternFill(start_color="F9C01A", end_color="F9C01A", fill_type="solid")
fill_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
fill_blue = PatternFill(start_color="A25555", end_color="A25555", fill_type="solid")

# Заполнение данных с чередующимися цветами строк
for row_num, row in enumerate(result.itertuples(index=False), 2):
    row_fill = fill_gray if row_num % 2 == 0 else fill_white
    for col_num, value in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        a = row[0]
        if row[0] in group_names:
            cell.fill = fill_green
        elif row[0] == '':
            cell.fill = fill_blue
        else:
            cell.fill = row_fill
        if isinstance(value, str):
            if value == "Н-Н":
                cell.value = 'Н'
                cell.fill = fill_red
            elif check_format(value):
                cell.fill = fill_yellow
            elif value == "Н":
                cell.fill = fill_red
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

# Замораживаем первую строку и колонку
ws.freeze_panes = "B2"

# Подгонка ширины колонок
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = max_length + 2

# Выбор пути для сохранения файла
save_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                              title="Сохранить файл как")
if save_path:
    wb.save(save_path)
    messagebox.showinfo("Сохранение файла", f"Файл успешно сохранен по пути:\n{save_path}")
else:
    print("Сохранение отменено.")
